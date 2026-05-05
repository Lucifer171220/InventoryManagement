from datetime import date, datetime, timedelta
from typing import List, Optional
from decimal import Decimal
from io import StringIO, BytesIO
from enum import Enum

from fastapi import APIRouter, Depends, HTTPException, Response, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, inspect
import csv

from app.database import get_db
from app.deps import get_current_active_user, require_role
from app.models import (
    User, InventoryItem, Warehouse, WarehouseInventory,
    Sale, SaleItem, PurchaseOrder, Supplier, Customer,
    StockTransfer, StockTransferItem, PurchaseOrderItem,
    Notification, AuditLog, HelpdeskConversation, HelpdeskMessage
)
from app.schemas import (
    InventoryReportFilter, InventoryReportItem, SalesReportFilter,
    InventoryAgingItem, ProfitMarginItem
)

# Import for PDF and Excel generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(prefix="/reports", tags=["Reports"])


TABLE_EXPORT_MODELS = {
    model.__tablename__: model
    for model in [
        Warehouse,
        InventoryItem,
        WarehouseInventory,
        StockTransfer,
        StockTransferItem,
        Supplier,
        PurchaseOrder,
        PurchaseOrderItem,
        Customer,
        Sale,
        SaleItem,
        Notification,
        AuditLog,
        HelpdeskConversation,
        HelpdeskMessage,
    ]
}


def _table_label(table_name: str) -> str:
    return table_name.replace("_", " ").title()


def _export_columns(model):
    return inspect(model).columns


def _format_export_value(value):
    if value is None:
        return ""
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, dict):
        return str(value)
    return value


def _get_export_model(table_name: str):
    model = TABLE_EXPORT_MODELS.get(table_name)
    if not model:
        raise HTTPException(status_code=404, detail="Report table not found")
    return model


@router.get("/table-exports")
def list_table_exports(
    current_user: User = Depends(get_current_active_user),
):
    """List reportable database tables. Users are intentionally excluded."""
    return [
        {"name": table_name, "label": _table_label(table_name)}
        for table_name in sorted(TABLE_EXPORT_MODELS)
    ]


@router.get("/table-exports/{table_name}/{format}")
def export_table_report(
    table_name: str,
    format: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export any allowed table as a report, excluding users."""
    model = _get_export_model(table_name)
    columns = list(_export_columns(model))
    headers = [column.name.replace("_", " ").title() for column in columns]
    rows = db.query(model).order_by(model.id).all()
    report_title = f"{_table_label(table_name)} Report"
    filename_base = f"{table_name}_report"

    if format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([_format_export_value(getattr(row, column.name)) for column in columns])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
        )

    if format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=500, detail="Excel generation not available. Install openpyxl.")

        wb = Workbook()
        ws = wb.active
        ws.title = _table_label(table_name)[:31]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            ws.append([_format_export_value(getattr(row, column.name)) for column in columns])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"}
        )

    if format == "pdf":
        if not HAS_REPORTLAB:
            raise HTTPException(status_code=500, detail="PDF generation not available. Install reportlab.")

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
        elements = []
        styles = getSampleStyleSheet()
        elements.append(Paragraph(report_title, styles["Heading1"]))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Rows: {len(rows)}", styles["Normal"]))
        elements.append(Spacer(1, 10))

        pdf_headers = headers[:8]
        pdf_columns = columns[:8]
        data = [pdf_headers]
        for row in rows[:500]:
            data.append([
                str(_format_export_value(getattr(row, column.name)))[:40]
                for column in pdf_columns
            ])

        if len(rows) > 500:
            elements.append(Paragraph("Showing first 500 rows in PDF export.", styles["Normal"]))
            elements.append(Spacer(1, 10))

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
        doc.build(elements)

        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.pdf"}
        )

    raise HTTPException(status_code=400, detail="Unsupported export format")


@router.post("/inventory")
def generate_inventory_report(
    filters: InventoryReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Generate inventory report with optional filters"""
    query = db.query(
        InventoryItem, WarehouseInventory, Warehouse, Supplier
    ).outerjoin(
        WarehouseInventory, InventoryItem.id == WarehouseInventory.item_id
    ).outerjoin(
        Warehouse, WarehouseInventory.warehouse_id == Warehouse.id
    ).outerjoin(
        Supplier, InventoryItem.supplier_id == Supplier.id
    )

    if filters.category:
        query = query.filter(InventoryItem.category == filters.category)
    if filters.warehouse_id:
        query = query.filter(WarehouseInventory.warehouse_id == filters.warehouse_id)
    if filters.low_stock_only:
        query = query.filter(InventoryItem.quantity <= InventoryItem.reorder_level)
    if filters.date_from:
        query = query.filter(InventoryItem.created_at >= filters.date_from)
    if filters.date_to:
        query = query.filter(InventoryItem.created_at <= filters.date_to)

    results = query.all()

    report_items = []
    for item, wi, warehouse, supplier in results:
        total_value = Decimal(str(item.quantity)) * item.unit_price
        report_items.append(InventoryReportItem(
            sku=item.sku,
            name=item.name,
            category=item.category,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total_value=total_value,
            warehouse_name=warehouse.name if warehouse else "N/A",
            supplier_name=supplier.name if supplier else "N/A"
        ))

    return {
        "generated_at": datetime.utcnow(),
        "total_items": len(report_items),
        "total_value": sum(item.total_value for item in report_items),
        "items": report_items
    }


@router.post("/inventory/export")
def export_inventory_report(
    format: str = "csv",  # csv, excel
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export inventory report to CSV"""
    items = db.query(InventoryItem, Supplier).outerjoin(
        Supplier, InventoryItem.supplier_id == Supplier.id
    ).filter(InventoryItem.is_active == True).all()

    output = StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        "SKU", "Name", "Category", "Subcategory", "Brand",
        "Quantity", "Unit Price", "Cost Price", "Total Value",
        "Reorder Level", "Supplier", "Warehouse", "Status"
    ])

    # Write data
    for item, supplier in items:
        total_value = Decimal(str(item.quantity)) * item.unit_price
        writer.writerow([
            item.sku,
            item.name,
            item.category,
            item.subcategory or "",
            item.brand or "",
            item.quantity,
            float(item.unit_price),
            float(item.cost_price),
            float(total_value),
            item.reorder_level,
            supplier.name if supplier else "N/A",
            item.warehouse_location or "N/A",
            "Active" if item.is_active else "Inactive"
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=inventory_report.csv"}
    )


@router.post("/sales")
def generate_sales_report(
    filters: SalesReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Generate sales report"""
    query = db.query(Sale, Customer).outerjoin(
        Customer, Sale.customer_id == Customer.id
    )

    query = query.filter(Sale.created_at >= filters.date_from)
    query = query.filter(Sale.created_at <= filters.date_to)

    if filters.customer_id:
        query = query.filter(Sale.customer_id == filters.customer_id)
    if filters.warehouse_id:
        query = query.filter(Sale.warehouse_id == filters.warehouse_id)

    sales = query.all()

    total_sales = len(sales)
    total_revenue = sum(sale.total_amount for sale, _ in sales)
    total_tax = sum(sale.tax_amount for sale, _ in sales)
    total_discount = sum(sale.discount_amount for sale, _ in sales)

    avg_order = total_revenue / total_sales if total_sales > 0 else Decimal("0")

    return {
        "period": {
            "from": filters.date_from,
            "to": filters.date_to
        },
        "summary": {
            "total_sales": total_sales,
            "total_revenue": total_revenue,
            "total_tax": total_tax,
            "total_discounts": total_discount,
            "average_order_value": avg_order
        },
        "sales": [
            {
                "sale_code": sale.sale_code,
                "date": sale.created_at,
                "customer": f"{customer.first_name} {customer.last_name}" if customer else "Walk-in",
                "total": sale.total_amount,
                "payment_method": sale.payment_method
            }
            for sale, customer in sales
        ]
    }


@router.get("/inventory-aging")
def get_inventory_aging(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get inventory aging report - items not sold in last X days"""
    from sqlalchemy import func

    ninety_days_ago = datetime.utcnow() - timedelta(days=90)

    items = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        InventoryItem.quantity > 0
    ).all()

    aging_items = []
    for item in items:
        # Get last sale date
        last_sale = db.query(func.max(Sale.created_at)).join(
            SaleItem
        ).filter(
            SaleItem.item_id == item.id,
            Sale.status == "completed"
        ).scalar()

        if last_sale:
            days_in_stock = (datetime.utcnow() - last_sale).days
            last_sale_date = last_sale
        else:
            days_in_stock = (datetime.utcnow() - item.created_at).days
            last_sale_date = None

        # Determine aging status
        if days_in_stock <= 30:
            status = "fresh"
        elif days_in_stock <= 60:
            status = "aging"
        elif days_in_stock <= 90:
            status = "stale"
        else:
            status = "dead"

        aging_items.append(InventoryAgingItem(
            sku=item.sku,
            name=item.name,
            category=item.category,
            quantity=item.quantity,
            days_in_stock=days_in_stock,
            last_sale_date=last_sale_date,
            aging_status=status
        ))

    return {
        "total_items": len(aging_items),
        "fresh": len([i for i in aging_items if i.aging_status == "fresh"]),
        "aging": len([i for i in aging_items if i.aging_status == "aging"]),
        "stale": len([i for i in aging_items if i.aging_status == "stale"]),
        "dead": len([i for i in aging_items if i.aging_status == "dead"]),
        "items": sorted(aging_items, key=lambda x: x.days_in_stock, reverse=True)
    }


@router.get("/profit-margins")
def get_profit_margins(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get profit margin analysis"""
    from sqlalchemy import func

    thirty_days_ago = datetime.utcnow() - timedelta(days=30)

    items = db.query(
        InventoryItem,
        func.sum(SaleItem.quantity).label("total_sold"),
        func.sum(SaleItem.total).label("total_revenue"),
        func.avg(SaleItem.unit_price).label("avg_sale_price")
    ).join(
        SaleItem, InventoryItem.id == SaleItem.item_id
    ).join(
        Sale, SaleItem.sale_id == Sale.id
    ).filter(
        Sale.created_at >= thirty_days_ago,
        Sale.status == "completed"
    ).group_by(InventoryItem.id).all()

    margin_items = []
    for item, total_sold, total_revenue, avg_sale_price in items:
        if total_sold and avg_sale_price:
            avg_sale = Decimal(str(avg_sale_price))
            cost = item.cost_price
            margin = avg_sale - cost
            margin_pct = (margin / avg_sale * 100) if avg_sale > 0 else 0

            margin_items.append(ProfitMarginItem(
                sku=item.sku,
                name=item.name,
                cost_price=cost,
                sale_price=avg_sale,
                margin_percent=float(margin_pct),
                total_sold=int(total_sold),
                total_profit=margin * int(total_sold)
            ))

    return {
        "period": "Last 30 days",
        "total_items_sold": sum(i.total_sold for i in margin_items),
        "total_profit": sum(i.total_profit for i in margin_items),
        "average_margin_percent": sum(i.margin_percent for i in margin_items) / len(margin_items) if margin_items else 0,
        "items": sorted(margin_items, key=lambda x: x.total_profit, reverse=True)
    }


@router.get("/supplier-performance")
def get_supplier_performance_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get supplier performance metrics"""
    from sqlalchemy import func

    suppliers = db.query(Supplier).filter(Supplier.is_active == True).all()

    performance = []
    for supplier in suppliers:
        # Total orders
        orders = db.query(PurchaseOrder).filter(
            PurchaseOrder.supplier_id == supplier.id,
            PurchaseOrder.status == "received"
        ).all()

        total_orders = len(orders)
        total_value = sum(o.total_amount for o in orders)

        # On-time delivery rate
        on_time = sum(1 for o in orders if o.actual_delivery and o.expected_delivery
                      and o.actual_delivery <= o.expected_delivery)
        on_time_rate = (on_time / total_orders * 100) if total_orders > 0 else 100

        # Average lead time
        avg_lead_time = None
        if orders:
            lead_times = [(o.actual_delivery - o.created_at).days for o in orders if o.actual_delivery]
            if lead_times:
                avg_lead_time = sum(lead_times) / len(lead_times)

        performance.append({
            "supplier_id": supplier.id,
            "supplier_name": supplier.name,
            "total_orders": total_orders,
            "total_value": total_value,
            "avg_lead_time": avg_lead_time,
            "on_time_delivery_rate": on_time_rate,
            "rating": float(supplier.rating)
        })

    return {
        "suppliers": sorted(performance, key=lambda x: x["total_value"], reverse=True)
    }


@router.get("/low-stock-export")
def export_low_stock(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export low stock items for reordering"""
    items = db.query(InventoryItem, Supplier).outerjoin(
        Supplier, InventoryItem.supplier_id == Supplier.id
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.quantity <= InventoryItem.reorder_level
    ).all()

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "SKU", "Name", "Current Qty", "Reorder Level", "Reorder Qty",
        "Supplier", "Supplier Email", "Unit Price", "Estimated Cost"
    ])

    for item, supplier in items:
        reorder_qty = max(item.reorder_quantity, item.reorder_level - item.quantity)
        estimated_cost = reorder_qty * item.cost_price
        writer.writerow([
            item.sku,
            item.name,
            item.quantity,
            item.reorder_level,
            reorder_qty,
            supplier.name if supplier else "N/A",
            supplier.email if supplier else "N/A",
            float(item.unit_price),
            float(estimated_cost)
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=reorder_list.csv"}
    )


# ============== PDF EXPORT ENDPOINTS ==============

@router.post("/inventory/export/pdf")
def export_inventory_report_pdf(
    filters: InventoryReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export inventory report to PDF"""
    if not HAS_REPORTLAB:
        raise HTTPException(status_code=500, detail="PDF generation not available. Install reportlab.")

    query = db.query(
        InventoryItem, WarehouseInventory, Warehouse, Supplier
    ).outerjoin(
        WarehouseInventory, InventoryItem.id == WarehouseInventory.item_id
    ).outerjoin(
        Warehouse, WarehouseInventory.warehouse_id == Warehouse.id
    ).outerjoin(
        Supplier, InventoryItem.supplier_id == Supplier.id
    )

    if filters.category:
        query = query.filter(InventoryItem.category == filters.category)
    if filters.warehouse_id:
        query = query.filter(WarehouseInventory.warehouse_id == filters.warehouse_id)
    if filters.low_stock_only:
        query = query.filter(InventoryItem.quantity <= InventoryItem.reorder_level)
    if filters.date_from:
        query = query.filter(InventoryItem.created_at >= filters.date_from)
    if filters.date_to:
        query = query.filter(InventoryItem.created_at <= filters.date_to)

    results = query.all()

    # Generate PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph("Inventory Report", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Generated info
    info = Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 12))

    # Summary
    total_items = len(results)
    total_value = sum(Decimal(str(item.quantity or 0)) * item.unit_price for item, _, _, _ in results)
    summary = Paragraph(f"Total Items: {total_items} | Total Value: ${float(total_value):,.2f}", styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 12))

    # Table data
    data = [["SKU", "Name", "Category", "Qty", "Unit Price", "Total Value", "Warehouse", "Supplier"]]
    for item, wi, warehouse, supplier in results:
        total_value_item = Decimal(str(item.quantity or 0)) * item.unit_price
        data.append([
            item.sku,
            item.name[:30] + "..." if len(item.name) > 30 else item.name,
            item.category,
            str(item.quantity),
            f"${float(item.unit_price):,.2f}",
            f"${float(total_value_item):,.2f}",
            warehouse.name if warehouse else "N/A",
            supplier.name if supplier else "N/A"
        ])

    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=inventory_report.pdf"}
    )


@router.post("/sales/export/pdf")
def export_sales_report_pdf(
    filters: SalesReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export sales report to PDF"""
    if not HAS_REPORTLAB:
        raise HTTPException(status_code=500, detail="PDF generation not available. Install reportlab.")

    query = db.query(Sale, Customer).outerjoin(
        Customer, Sale.customer_id == Customer.id
    )

    query = query.filter(Sale.created_at >= filters.date_from)
    query = query.filter(Sale.created_at <= filters.date_to)

    if filters.customer_id:
        query = query.filter(Sale.customer_id == filters.customer_id)
    if filters.warehouse_id:
        query = query.filter(Sale.warehouse_id == filters.warehouse_id)

    sales = query.all()

    total_sales = len(sales)
    total_revenue = sum(sale.total_amount for sale, _ in sales)
    total_tax = sum(sale.tax_amount for sale, _ in sales)
    total_discount = sum(sale.discount_amount for sale, _ in sales)
    avg_order = total_revenue / total_sales if total_sales > 0 else Decimal("0")

    # Generate PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph("Sales Report", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Period info
    period = Paragraph(f"Period: {filters.date_from.strftime('%Y-%m-%d')} to {filters.date_to.strftime('%Y-%m-%d')}", styles['Normal'])
    elements.append(period)
    elements.append(Spacer(1, 12))

    # Summary
    summary_data = [
        ["Total Sales", "Total Revenue", "Total Tax", "Total Discounts", "Avg Order Value"],
        [str(total_sales), f"${float(total_revenue):,.2f}", f"${float(total_tax):,.2f}",
         f"${float(total_discount):,.2f}", f"${float(avg_order):,.2f}"]
    ]
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#D1FAE5')),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Sales table
    if sales:
        data = [["Sale Code", "Date", "Customer", "Payment Method", "Total"]]
        for sale, customer in sales:
            data.append([
                sale.sale_code,
                sale.created_at.strftime('%Y-%m-%d'),
                f"{customer.first_name} {customer.last_name}" if customer else "Walk-in",
                sale.payment_method,
                f"${float(sale.total_amount):,.2f}"
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sales_report.pdf"}
    )


# ============== EXCEL/XLSX EXPORT ENDPOINTS ==============

@router.post("/inventory/export/excel")
def export_inventory_report_excel(
    filters: InventoryReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export inventory report to Excel/XLSX"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="Excel generation not available. Install openpyxl.")

    query = db.query(
        InventoryItem, WarehouseInventory, Warehouse, Supplier
    ).outerjoin(
        WarehouseInventory, InventoryItem.id == WarehouseInventory.item_id
    ).outerjoin(
        Warehouse, WarehouseInventory.warehouse_id == Warehouse.id
    ).outerjoin(
        Supplier, InventoryItem.supplier_id == Supplier.id
    )

    if filters.category:
        query = query.filter(InventoryItem.category == filters.category)
    if filters.warehouse_id:
        query = query.filter(WarehouseInventory.warehouse_id == filters.warehouse_id)
    if filters.low_stock_only:
        query = query.filter(InventoryItem.quantity <= InventoryItem.reorder_level)
    if filters.date_from:
        query = query.filter(InventoryItem.created_at >= filters.date_from)
    if filters.date_to:
        query = query.filter(InventoryItem.created_at <= filters.date_to)

    results = query.all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    # Header row
    headers = ["SKU", "Name", "Category", "Quantity", "Unit Price", "Cost Price", "Total Value",
               "Reorder Level", "Warehouse", "Supplier", "Status"]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_value = Decimal("0")
    for item, wi, warehouse, supplier in results:
        item_total = Decimal(str(item.quantity or 0)) * item.unit_price
        total_value += item_total
        ws.append([
            item.sku,
            item.name,
            item.category,
            item.quantity,
            float(item.unit_price),
            float(item.cost_price),
            float(item_total),
            item.reorder_level,
            warehouse.name if warehouse else "N/A",
            supplier.name if supplier else "N/A",
            "Active" if item.is_active else "Inactive"
        ])

    # Summary row
    ws.append([])
    ws.append(["Total Items:", len(results), "", "", "", "", float(total_value)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"}
    )


@router.post("/sales/export/excel")
def export_sales_report_excel(
    filters: SalesReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export sales report to Excel/XLSX"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="Excel generation not available. Install openpyxl.")

    query = db.query(Sale, Customer).outerjoin(
        Customer, Sale.customer_id == Customer.id
    )

    query = query.filter(Sale.created_at >= filters.date_from)
    query = query.filter(Sale.created_at <= filters.date_to)

    if filters.customer_id:
        query = query.filter(Sale.customer_id == filters.customer_id)
    if filters.warehouse_id:
        query = query.filter(Sale.warehouse_id == filters.warehouse_id)

    sales = query.all()

    # Create workbook
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    total_sales = len(sales)
    total_revenue = sum(sale.total_amount for sale, _ in sales)
    total_tax = sum(sale.tax_amount for sale, _ in sales)
    total_discount = sum(sale.discount_amount for sale, _ in sales)
    avg_order = total_revenue / total_sales if total_sales > 0 else Decimal("0")

    ws_summary.append(["Sales Report Summary"])
    ws_summary.append([])
    ws_summary.append(["Period", f"{filters.date_from.strftime('%Y-%m-%d')} to {filters.date_to.strftime('%Y-%m-%d')}"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Sales", total_sales])
    ws_summary.append(["Total Revenue", float(total_revenue)])
    ws_summary.append(["Total Tax", float(total_tax)])
    ws_summary.append(["Total Discounts", float(total_discount)])
    ws_summary.append(["Average Order Value", float(avg_order)])

    # Style summary
    ws_summary['A1'].font = Font(bold=True, size=14)
    for row in ws_summary['A5:B10']:
        for cell in row:
            if cell.row == 5:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    # Sales detail sheet
    ws_detail = wb.create_sheet("Sales Detail")
    ws_detail.append(["Sale Code", "Date", "Customer", "Subtotal", "Tax", "Discount", "Total", "Payment Method"])

    # Style header
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for sale, customer in sales:
        ws_detail.append([
            sale.sale_code,
            sale.created_at.strftime('%Y-%m-%d %H:%M'),
            f"{customer.first_name} {customer.last_name}" if customer else "Walk-in",
            float(sale.subtotal),
            float(sale.tax_amount),
            float(sale.discount_amount),
            float(sale.total_amount),
            sale.payment_method
        ])

    # Auto-adjust column widths
    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
    )
