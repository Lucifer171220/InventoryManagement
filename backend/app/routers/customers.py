from datetime import datetime
from typing import List, Optional
from decimal import Decimal
from io import StringIO, BytesIO
import csv
import logging

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, desc

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from app.database import get_db
from app.deps import get_current_active_user, require_role
from app.models import User, Customer, Sale
from app.schemas import (
    CustomerCreate, CustomerUpdate, CustomerResponse, CustomerPurchaseHistory
)

router = APIRouter(prefix="/customers", tags=["Customers"])
logger = logging.getLogger(__name__)


@router.get("/", response_model=List[CustomerResponse])
def list_customers(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    loyalty_tier: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    try:
        query = db.query(Customer)

        if search:
            search_term = f"%{search}%"
            query = query.filter(
                Customer.first_name.ilike(search_term) |
                Customer.last_name.ilike(search_term) |
                Customer.email.ilike(search_term) |
                Customer.phone.ilike(search_term) |
                Customer.customer_code.ilike(search_term)
            )

        if loyalty_tier:
            query = query.filter(Customer.loyalty_tier == loyalty_tier)

        customers = query.order_by(Customer.id).offset(skip).limit(limit).all()

        result = []
        for c in customers:
            # Calculate total purchases and orders
            purchase_stats = db.query(
                func.count(Sale.id).label("order_count"),
                func.sum(Sale.total_amount).label("total_purchases")
            ).filter(
                Sale.customer_id == c.id,
                Sale.status == "completed"
            ).first()

            result.append({
                "id": c.id,
                "customer_code": c.customer_code,
                "first_name": c.first_name,
                "last_name": c.last_name,
                "email": c.email,
                "phone": c.phone,
                "address": c.address,
                "city": c.city,
                "state": c.state,
                "postal_code": c.postal_code,
                "country": c.country,
                "date_of_birth": c.date_of_birth,
                "loyalty_points": c.loyalty_points,
                "loyalty_tier": c.loyalty_tier,
                "is_active": c.is_active,
                "notes": c.notes,
                "total_purchases": purchase_stats.total_purchases or Decimal("0"),
                "total_orders": purchase_stats.order_count or 0,
                "created_at": c.created_at
            })

        return result
    except Exception as e:
        logger.error(f"Error in list_customers: {type(e).__name__}: {e}")
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@router.post("/", response_model=CustomerResponse, status_code=status.HTTP_201_CREATED)
def create_customer(
    customer: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    # Generate customer code if not provided
    if not customer.customer_code:
        last_customer = db.query(Customer).order_by(desc(Customer.id)).first()
        next_id = (last_customer.id + 1) if last_customer else 1
        customer_code = f"CUST{next_id:06d}"
    else:
        # Check if code exists
        existing = db.query(Customer).filter(Customer.customer_code == customer.customer_code).first()
        if existing:
            raise HTTPException(status_code=400, detail="Customer code already exists")
        customer_code = customer.customer_code

    db_customer = Customer(**customer.model_dump(exclude={"customer_code"}), customer_code=customer_code)
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return get_customer(db_customer.id, db=db, current_user=current_user)



@router.get("/{customer_id}", response_model=CustomerResponse)
def get_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Calculate purchase stats
    purchase_stats = db.query(
        func.count(Sale.id).label("order_count"),
        func.sum(Sale.total_amount).label("total_purchases")
    ).filter(
        Sale.customer_id == customer_id,
        Sale.status == "completed"
    ).first()

    return {
        "id": customer.id,
        "customer_code": customer.customer_code,
        "first_name": customer.first_name,
        "last_name": customer.last_name,
        "email": customer.email,
        "phone": customer.phone,
        "address": customer.address,
        "city": customer.city,
        "state": customer.state,
        "postal_code": customer.postal_code,
        "country": customer.country,
        "date_of_birth": customer.date_of_birth,
        "loyalty_points": customer.loyalty_points,
        "loyalty_tier": customer.loyalty_tier,
        "is_active": customer.is_active,
        "notes": customer.notes,
        "total_purchases": purchase_stats.total_purchases or Decimal("0"),
        "total_orders": purchase_stats.order_count or 0,
        "created_at": customer.created_at
    }


@router.put("/{customer_id}", response_model=CustomerResponse)
def update_customer(
    customer_id: int,
    customer_update: CustomerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    for field, value in customer_update.model_dump(exclude_unset=True).items():
        setattr(customer, field, value)

    db.commit()
    db.refresh(customer)
    return customer


@router.delete("/{customer_id}")
def delete_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    # Soft delete
    customer.is_active = False
    db.commit()
    return {"message": "Customer deactivated successfully"}


@router.get("/{customer_id}/purchase-history")
def get_purchase_history(
    customer_id: int,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    sales = db.query(Sale).filter(
        Sale.customer_id == customer_id
    ).order_by(desc(Sale.created_at)).offset(skip).limit(limit).all()

    return [
        CustomerPurchaseHistory(
            sale_id=sale.id,
            sale_code=sale.sale_code,
            sale_date=sale.created_at,
            total_amount=sale.total_amount,
            items_count=len(sale.items)
        )
        for sale in sales
    ]


@router.post("/{customer_id}/add-loyalty-points")
def add_loyalty_points(
    customer_id: int,
    points: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    customer = db.query(Customer).filter(Customer.id == customer_id).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    customer.loyalty_points += points

    # Update tier based on points
    if customer.loyalty_points >= 10000:
        customer.loyalty_tier = "platinum"
    elif customer.loyalty_points >= 5000:
        customer.loyalty_tier = "gold"
    elif customer.loyalty_points >= 1000:
        customer.loyalty_tier = "silver"
    else:
        customer.loyalty_tier = "bronze"

    db.commit()
    return {
        "message": "Loyalty points added successfully",
        "current_points": customer.loyalty_points,
        "tier": customer.loyalty_tier
    }


# ============== EXPORT ENDPOINTS ==============

@router.get("/export/csv")
def export_customers_csv(
    search: Optional[str] = None,
    loyalty_tier: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export customers to CSV"""
    query = db.query(Customer)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            Customer.first_name.ilike(search_term) |
            Customer.last_name.ilike(search_term) |
            Customer.email.ilike(search_term) |
            Customer.phone.ilike(search_term)
        )

    if loyalty_tier:
        query = query.filter(Customer.loyalty_tier == loyalty_tier)

    customers = query.all()

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Customer Code", "First Name", "Last Name", "Email", "Phone",
        "Address", "City", "State", "Postal Code", "Country",
        "Loyalty Points", "Loyalty Tier", "Is Active", "Created At"
    ])

    for c in customers:
        writer.writerow([
            c.customer_code or "",
            c.first_name,
            c.last_name,
            c.email or "",
            c.phone or "",
            c.address or "",
            c.city or "",
            c.state or "",
            c.postal_code or "",
            c.country or "",
            c.loyalty_points,
            c.loyalty_tier,
            "Yes" if c.is_active else "No",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else ""
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=customers_export.csv"}
    )


@router.get("/export/excel")
def export_customers_excel(
    search: Optional[str] = None,
    loyalty_tier: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export customers to Excel/XLSX"""
    if not HAS_OPENPYXL:
        raise HTTPException(status_code=500, detail="Excel generation not available. Install openpyxl.")

    query = db.query(Customer)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            Customer.first_name.ilike(search_term) |
            Customer.last_name.ilike(search_term) |
            Customer.email.ilike(search_term) |
            Customer.phone.ilike(search_term)
        )

    if loyalty_tier:
        query = query.filter(Customer.loyalty_tier == loyalty_tier)

    customers = query.all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header row
    headers = [
        "Customer Code", "First Name", "Last Name", "Email", "Phone",
        "Address", "City", "State", "Postal Code", "Country",
        "Loyalty Points", "Loyalty Tier", "Is Active", "Created At"
    ]
    ws.append(headers)

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for c in customers:
        ws.append([
            c.customer_code or "",
            c.first_name,
            c.last_name,
            c.email or "",
            c.phone or "",
            c.address or "",
            c.city or "",
            c.state or "",
            c.postal_code or "",
            c.country or "",
            c.loyalty_points,
            c.loyalty_tier,
            "Yes" if c.is_active else "No",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else ""
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers_export.xlsx"}
    )


@router.get("/export/pdf")
def export_customers_pdf(
    search: Optional[str] = None,
    loyalty_tier: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("manager", "moderator")),
):
    """Export customers to PDF"""
    if not HAS_REPORTLAB:
        raise HTTPException(status_code=500, detail="PDF generation not available. Install reportlab.")

    query = db.query(Customer)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            Customer.first_name.ilike(search_term) |
            Customer.last_name.ilike(search_term) |
            Customer.email.ilike(search_term) |
            Customer.phone.ilike(search_term)
        )

    if loyalty_tier:
        query = query.filter(Customer.loyalty_tier == loyalty_tier)

    customers = query.all()

    # Generate PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph("Customers Report", styles['Heading1'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Generated info
    info = Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 12))

    # Summary
    summary = Paragraph(f"Total Customers: {len(customers)}", styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 12))

    # Table data
    data = [[
        "Customer Code", "Name", "Email", "Phone",
        "Loyalty Points", "Tier", "Status"
    ]]

    for c in customers:
        data.append([
            c.customer_code or "N/A",
            f"{c.first_name} {c.last_name}",
            c.email or "N/A",
            c.phone or "N/A",
            str(c.loyalty_points),
            c.loyalty_tier,
            "Active" if c.is_active else "Inactive"
        ])

    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=customers_report.pdf"}
    )
